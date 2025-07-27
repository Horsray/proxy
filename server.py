#
# Enhanced User management server for Huiying Proxy
# 绘影用户管理系统
# 本文件为绘影代理用户管理系统主服务端代码，负责用户、产品、台账、审批等核心业务逻辑。
# 各部分函数均有详细注释，说明用途、流程、交互及异常处理。
# 导入标准库及所需第三方库
import os
import json
import inspect
import argparse
import requests
from datetime import datetime
from io import BytesIO
from functools import wraps

# 导入Flask及相关工具
from flask import Flask, render_template, request, redirect, url_for, session, send_file, jsonify
from werkzeug.utils import secure_filename

# Excel文件处理库
from openpyxl import Workbook, load_workbook
# 变量定义
# 基础目录及数据文件路径
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
USERS_FILE = os.path.join(BASE_DIR, 'users.json')          # 用户数据文件
LEDGER_FILE = os.path.join(BASE_DIR, 'ledger.json')        # 台账数据文件
PRODUCTS_FILE = os.path.join(BASE_DIR, 'products.json')    # 产品数据文件
APPLICATIONS_FILE = os.path.join(BASE_DIR, 'applications.json') # 审批数据文件

# Flask应用初始化及密钥设置
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'huiying-secret')  # 用户身份码，生产环境请设置环境变量


def get_location_from_ip(ip_address):
    """根据IP地址获取地理位置信息

    用途：通过公网API获取IP归属地信息（国家-省份-城市），失败时返回原始IP。
    交互：对外部ip-api.com发起HTTP请求。
    异常：任何异常（如网络、解析等）均忽略并返回原始IP。
    """
    if not ip_address:
        return ''
    try:
        resp = requests.get(
            f"http://ip-api.com/json/{ip_address}?lang=zh-CN",
            timeout=3,
        )
        if resp.status_code == 200:
            data = resp.json()
            if data.get('status') == 'success':
                country = data.get('country', '')
                region = data.get('regionName', '')
                city = data.get('city', '')
                location_parts = [p for p in [country, region, city] if p]
                if location_parts:
                    return '-'.join(location_parts)
    except Exception:
        pass
    return ip_address


def get_client_ip():
    """获取客户端真实IP地址
    用途：兼容代理环境下获取用户真实IP，优先X-Forwarded-For，再X-Real-IP，否则取remote_addr。
    """
    if request.headers.get('X-Forwarded-For'):
        return request.headers.get('X-Forwarded-For').split(',')[0].strip()
    elif request.headers.get('X-Real-IP'):
        return request.headers.get('X-Real-IP')
    else:
        return request.remote_addr


def generate_user_id(users: dict) -> str:
    """
    生成唯一用户编号。编号格式为当前时间戳+三位序号，确保唯一性。
    参数:
        users: 当前所有用户字典，用于检测已存在编号。
    返回:
        新的用户编号字符串。
    """
    ts = datetime.now().strftime('%Y%m%d%H%M%S')
    seq = 1
    for u in users.values():
        uid = str(u.get('user_id', ''))
        if uid.startswith(ts):
            try:
                num = int(uid[len(ts):])
                if num >= seq:
                    seq = num + 1
            except Exception:
                continue
    return f"{ts}{seq:03d}"


def load_users() -> dict:
    """
    加载用户数据文件，若缺省字段则自动补全，必要时回写文件。
    用途：全局用户数据读取与字段规范化。
    异常：文件不存在或格式错误时返回空字典。
    """
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            try:
                users = json.load(f).get('users', {})
                # 确保每个用户字段齐全
                updated = False
                for name, info in users.items():
                    info.setdefault('nickname', '')
                    info.setdefault('enabled', True)
                    info.setdefault('source', 'add')
                    info.setdefault('price', 0)
                    info.setdefault('ip_address', '')
                    info.setdefault('location', '')
                    info.setdefault('remark', '')
                    info.setdefault('is_agent', False)
                    info.setdefault('owner', '')
                    info.setdefault('forsale', False)
                    if 'user_id' not in info:
                        info['user_id'] = generate_user_id(users)
                        updated = True
                    # 若有IP但无地理位置则补全
                    if info.get('ip_address') and not info.get('location'):
                        info['location'] = get_location_from_ip(info['ip_address'])
                        updated = True
                if updated:
                    save_users(users)
                return users
            except Exception:
                return {}
    return {}


def save_users(users: dict) -> None:
    """
    保存用户数据到文件。
    用途：对用户信息的持久化。
    """
    with open(USERS_FILE, 'w', encoding='utf-8') as f:
        json.dump({'users': users}, f, indent=4, ensure_ascii=False)


def load_ledger() -> list:
    """
    加载台账记录列表，自动补全role字段。
    用途：用于收入、销售等统计与显示。
    异常：文件不存在/损坏时返回空列表。
    """
    if os.path.exists(LEDGER_FILE):
        with open(LEDGER_FILE, 'r', encoding='utf-8') as f:
            try:
                records = json.load(f).get('records', [])
                for r in records:
                    r.setdefault('role', 'admin')
                return records
            except Exception:
                return []
    return []


def save_ledger(records: list) -> None:
    """
    保存台账记录到文件。
    """
    with open(LEDGER_FILE, 'w', encoding='utf-8') as f:
        json.dump({'records': records}, f, indent=4, ensure_ascii=False)


def load_products() -> dict:
    """
    加载产品信息字典，补全缺省字段（价格、默认标志）。
    用途：产品管理与下拉选择。
    异常：文件不存在/损坏时返回空字典。
    """
    if os.path.exists(PRODUCTS_FILE):
        with open(PRODUCTS_FILE, 'r', encoding='utf-8') as f:
            try:
                products = json.load(f).get('products', {})
                for p in products.values():
                    p.setdefault('price', 0)
                    p.setdefault('default', False)
                return products
            except Exception:
                return {}
    return {}


def save_products(products: dict) -> None:
    """
    保存产品信息到文件。
    """
    with open(PRODUCTS_FILE, 'w', encoding='utf-8') as f:
        json.dump({'products': products}, f, indent=4, ensure_ascii=False)


def load_applications() -> list:
    """
    加载代理批量申请记录列表。
    用途：审批管理。
    异常：文件不存在/损坏时返回空列表。
    """
    if os.path.exists(APPLICATIONS_FILE):
        with open(APPLICATIONS_FILE, 'r', encoding='utf-8') as f:
            try:
                return json.load(f).get('apps', [])
            except Exception:
                return []
    return []


def save_applications(apps: list) -> None:
    """
    保存代理批量申请记录到文件。
    """
    with open(APPLICATIONS_FILE, 'w', encoding='utf-8') as f:
        json.dump({'apps': apps}, f, indent=4, ensure_ascii=False)


@app.context_processor
def inject_counts():
    """
    模板上下文处理器：为模板提供待审批/待申请数量。
    用途：页面角标、提示等。
    """
    apps = load_applications()
    pending_admin = sum(1 for a in apps if a.get('status') == 'pending')
    pending_agent = 0
    if session.get('agent'):
        pending_agent = sum(
            1 for a in apps
            if a.get('agent') == session.get('agent') and a.get('status') == 'pending'
        )
    return dict(
        pending_approve_count=pending_admin,
        pending_apply_count=pending_agent,
    )


def admin_required(f):
    """
    装饰器：要求管理员身份访问。未登录重定向到登录页。
    """
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('admin'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper


def agent_required(f):
    """
    装饰器：要求销售代理身份访问。未登录重定向到登录页。
    """
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('agent'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper


@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    登录页面处理（管理员与代理）。
    用途：身份验证、会话初始化、IP及地理位置记录。
    交互：读取用户数据，写入登录信息。
    异常：密码错误或用户不存在时返回错误提示。
    """
    # 清除上次登录状态，避免角色混淆
    session.pop('admin', None)
    session.pop('agent', None)

    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        users = load_users()
        user = users.get(username)
        if user and user.get('password') == password:
            # 登录成功，记录登录时间和来源IP
            client_ip = get_client_ip()
            user['last_login'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            user['ip_address'] = client_ip
            user['location'] = get_location_from_ip(client_ip)
            users[username] = user
            save_users(users)
            if user.get('is_admin'):
                session['admin'] = username
                return redirect(url_for('user_list'))
            if user.get('is_agent'):
                session['agent'] = username
                return redirect(url_for('agent_users'))
        # 登录失败
        return render_template('login.html', error='登录失败')
    return render_template('login.html')


@app.route('/logout')
def logout():
    """
    登出操作，清除登录状态。
    """
    session.pop('admin', None)
    session.pop('agent', None)
    return redirect(url_for('login'))


@app.route('/')
@admin_required
def index():
    """
    首页重定向到用户列表，仅管理员可访问。
    """
    return redirect(url_for('user_list'))


@app.route('/bulk')
@admin_required
def bulk_manage():
    """
    批量操作管理页面，显示最近批量创建的账户信息。
    用途：批量导出、回显等。
    交互：读取会话中的批量账户列表。
    """
    accounts = session.get('bulk_accounts')
    info = session.get('bulk_info')
    products = load_products()
    page = int(request.args.get('page', 1))
    per_page = 20
    if accounts:
        start = (page - 1) * per_page
        page_accounts = accounts[start:start + per_page]
    else:
        page_accounts = None
    return render_template(
        'bulk.html', accounts=page_accounts, products=products,
        info=info, page=page, per_page=per_page,
        total=len(accounts) if accounts else 0
    )


@app.route('/users/random')
@admin_required
def random_account():
    """
    随机生成一个未占用的用户名及密码。
    用途：前端快速生成新用户账号。
    """
    users = load_users()
    while True:
        uname = f"huiying{os.urandom(4).hex()}"[:12]
        if uname not in users:
            break
    pwd = os.urandom(4).hex()
    return jsonify({'username': uname, 'password': pwd})


@app.route('/bulk/export')
@admin_required
def bulk_export():
    """
    导出最近一次批量创建的账户为Excel文件。
    用途：管理员批量导出分发。
    交互：从session获取账户列表，生成Excel并下载。
    """
    accounts = session.get('bulk_accounts')
    if not accounts:
        return redirect(url_for('bulk_manage'))
    wb = Workbook()
    ws = wb.active
    ws.append(['用户名', '密码'])
    for acc in accounts:
        ws.append([acc['username'], acc['password']])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    # 兼容不同Flask版本的下载参数
    filename = 'bulk_accounts.xlsx'
    kwargs = {}
    if 'download_name' in inspect.signature(send_file).parameters:
        kwargs['download_name'] = filename
    else:
        kwargs['attachment_filename'] = filename
    response = send_file(
        bio,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        **kwargs
    )
    return response


@app.route('/users')
@admin_required
def user_list():
    """
    用户列表页面，支持多条件筛选、分页、排序。
    用途：管理员管理所有用户。
    交互：前端参数过滤，展示用户及产品信息。
    """
    query = request.args.get('q', '')
    source = request.args.get('source', '')
    status = request.args.get('status', '')
    sale = request.args.get('sale', '')
    sort = request.args.get('sort', 'desc')
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    page = int(request.args.get('page', 1))
    per_page = max(int(request.args.get('per_page', 10)), 1)
    users = load_users()
    # 多条件筛选
    if query:
        users = {k: v for k, v in users.items() if query.lower() in k.lower()}
    if source:
        users = {k: v for k, v in users.items() if v.get('source') == source}
    if status:
        flag = status == 'enabled'
        users = {k: v for k, v in users.items() if v.get('enabled', True) == flag}
    if sale:
        flag = sale == 'forsale'
        users = {k: v for k, v in users.items() if v.get('forsale', False) == flag}
    if start:
        users = {k: v for k, v in users.items() if v.get('created_at', '') >= start}
    if end:
        users = {k: v for k, v in users.items() if v.get('created_at', '') <= end}
    items = list(users.items())
    # 管理员用户优先显示
    admins = [i for i in items if i[1].get('is_admin')]
    others = [i for i in items if not i[1].get('is_admin')]
    others.sort(key=lambda x: x[1].get('created_at', ''), reverse=(sort != 'asc'))
    items = admins + others
    total = len(items)
    page_items = items[(page - 1) * per_page: page * per_page]
    products = load_products()
    return render_template(
        'users.html', users=dict(page_items), total=total,
        page=page, per_page=per_page, query=query, source=source,
        status=status, sale=sale, sort=sort, start=start, end=end,
        products=products
    )


@app.route('/users/add', methods=['POST'])
@admin_required
def add_user():
    """
    添加新用户（管理员操作）。
    用途：表单提交新用户，写入用户数据并记录台账。
    交互：重名检查、字段补全。
    """
    username = request.form.get('username')
    password = request.form.get('password')
    nickname = request.form.get('nickname', '')
    is_admin = bool(request.form.get('is_admin'))
    is_agent = bool(request.form.get('is_agent'))
    price = float(request.form.get('price') or 0)
    product = request.form.get('product', '')
    if not username or not password:
        return redirect(url_for('user_list'))
    users = load_users()
    if username in users:
        return redirect(url_for('user_list'))
    users[username] = {
        'user_id': generate_user_id(users),
        'password': password,
        'nickname': nickname,
        'is_admin': is_admin,
        'is_agent': is_agent,
        'enabled': True,
        'source': 'add',
        'price': price,
        'product': product,
        'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'last_login': None,
        'ip_address': '',
        'location': '',
        'remark': ''
    }
    save_users(users)
    # ledger
    records = load_ledger()
    records.append({
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'admin': session.get('admin'),
        'role': 'admin',
        'product': product,
        'price': price,
        'count': 1,
        'revenue': price
    })
    save_ledger(records)
    return redirect(url_for('user_list'))


@app.route('/users/<name>/delete', methods=['POST'])
@admin_required
def delete_user(name):
    """
    删除指定用户（管理员操作）。
    用途：用户数据清理。
    """
    users = load_users()
    if name in users:
        users.pop(name)
        save_users(users)
    return redirect(url_for('user_list'))


@app.route('/users/<name>/toggle', methods=['POST'])
def toggle_user(name):
    """
    启用/禁用用户（管理员或所属代理可操作）。
    用途：账号管控，支持AJAX和表单。
    安全：仅管理员horsray或代理本人可操作。
    """
    users = load_users()
    user = users.get(name)
    permitted = False
    if session.get('admin') == 'horsray':
        permitted = True
    elif session.get('agent') and user and user.get('owner') == session.get('agent'):
        permitted = True
    if not permitted:
        return redirect(url_for('login'))
    if user:
        user['enabled'] = not user.get('enabled', True)
        save_users(users)
        if request.is_json or request.headers.get('Content-Type') == 'application/json':
            return jsonify({'success': True, 'enabled': user['enabled']})
    if request.is_json or request.headers.get('Content-Type') == 'application/json':
        return jsonify({'success': False}), 404
    return redirect(url_for('user_list') if session.get('admin') else url_for('agent_users'))


@app.route('/sales/users/<name>/sold', methods=['POST'])
@agent_required
def mark_sold(name):
    """
    标记代理名下某账号为已售出。
    用途：代理销售记录台账。
    交互：仅代理本人且账号处于待售状态可操作。
    """
    users = load_users()
    current = session.get('agent')
    if name in users and users[name].get('owner') == current and users[name].get('forsale'):
        users[name]['forsale'] = False
        save_users(users)
        # add ledger record when item sold
        records = load_ledger()
        price = users[name].get('price', 0)
        records.append({
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'admin': current,
            'role': 'agent',
            'product': users[name].get('product', ''),
            'price': price,
            'count': 1,
            'revenue': price
        })
        save_ledger(records)
        if request.is_json or request.headers.get('Accept') == 'application/json':
            return jsonify({'success': True})
    if request.is_json or request.headers.get('Accept') == 'application/json':
        return jsonify({'success': False}), 404
    return redirect(url_for('agent_users'))


@app.route('/sales/batch_sold', methods=['POST'])
@agent_required
def batch_sold():
    """
    批量标记代理名下账号为已售出。
    用途：代理批量销售，台账同步记录。
    """
    names = request.form.getlist('names')
    users = load_users()
    current = session.get('agent')
    sold_any = False
    for name in names:
        if name in users and users[name].get('owner') == current and users[name].get('forsale'):
            users[name]['forsale'] = False
            price = users[name].get('price', 0)
            records = load_ledger()
            records.append({
                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'admin': current,
                'role': 'agent',
                'product': users[name].get('product', ''),
                'price': price,
                'count': 1,
                'revenue': price
            })
            save_ledger(records)
            sold_any = True
    save_users(users)
    return redirect(url_for('agent_users'))


@app.route('/sales/users/<name>/update', methods=['POST'])
@agent_required
def agent_update_user(name):
    """
    代理修改自己名下账号信息（支持AJAX更新备注）。
    用途：代理自助管理。
    安全：仅限本人；支持用户名变更、密码、昵称、产品、备注等。
    """
    users = load_users()
    current = session.get('agent')
    user = users.get(name)
    if not user or user.get('owner') != current:
        if request.is_json:
            return jsonify({'success': False}), 404
        return redirect(url_for('agent_users'))
    if request.is_json:
        data = request.get_json(silent=True) or {}
        remark = data.get('remark')
        if remark is not None:
            user['remark'] = remark
            save_users(users)
            return jsonify({'success': True})
        return jsonify({'success': False}), 400
    new_name = request.form.get('username')
    password = request.form.get('password')
    nickname = request.form.get('nickname')
    enabled = bool(request.form.get('enabled'))
    product = request.form.get('product')
    remark = request.form.get('remark', '')
    if new_name and new_name != name and new_name not in users:
        users[new_name] = user
        users.pop(name)
        name = new_name
        user = users[name]
    if password:
        user['password'] = password
    if nickname is not None:
        user['nickname'] = nickname
    user['enabled'] = enabled
    if product is not None:
        user['product'] = product
    user['remark'] = remark
    save_users(users)
    return redirect(url_for('agent_users'))


@app.route('/users/batch_action', methods=['POST'])
@admin_required
def batch_action():
    """
    管理员批量操作用户（删除、启用、禁用）。
    用途：多选批量管理。
    """
    action = request.form.get('action')
    names = request.form.getlist('names')
    users = load_users()
    for name in names:
        if name not in users:
            continue
        if action == 'delete':
            users.pop(name)
        elif action == 'enable':
            users[name]['enabled'] = True
        elif action == 'disable':
            users[name]['enabled'] = False
    save_users(users)
    return redirect(url_for('user_list'))


@app.route('/sales/batch_action', methods=['POST'])
@agent_required
def agent_batch_action():
    """
    代理批量操作自己名下用户（启用、禁用、标记已售）。
    用途：代理自助多选管理。
    """
    action = request.form.get('action')
    names = request.form.getlist('names')
    users = load_users()
    current = session.get('agent')
    for name in names:
        if name not in users or users[name].get('owner') != current:
            continue
        if action == 'enable':
            users[name]['enabled'] = True
        elif action == 'disable':
            users[name]['enabled'] = False
        elif action == 'sold' and users[name].get('forsale'):
            users[name]['forsale'] = False
            price = users[name].get('price', 0)
            records = load_ledger()
            records.append({
                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'admin': current,
                'role': 'agent',
                'product': users[name].get('product', ''),
                'price': price,
                'count': 1,
                'revenue': price
            })
            save_ledger(records)
    save_users(users)
    return redirect(url_for('agent_users'))


@app.route('/users/<name>/update', methods=['POST'])
@admin_required
def update_user(name):
    """
    管理员更新用户信息（支持AJAX备注更新）。
    用途：支持表单与AJAX两种方式。
    """
    users = load_users()
    if request.is_json:
        data = request.get_json(silent=True) or {}
        remark = data.get('remark')
        if name in users and remark is not None:
            users[name]['remark'] = remark
            save_users(users)
            return jsonify({'success': True})
        return jsonify({'success': False}), 404

    new_name = request.form.get('username')
    password = request.form.get('password')
    nickname = request.form.get('nickname')
    is_admin = bool(request.form.get('is_admin'))
    is_agent = bool(request.form.get('is_agent'))
    enabled = bool(request.form.get('enabled'))
    product = request.form.get('product')
    remark = request.form.get('remark', '')
    if name in users:
        user = users[name]
        if new_name and new_name != name:
            users[new_name] = user
            users.pop(name)
            name = new_name
        if password:
            users[name]['password'] = password
        if nickname is not None:
            users[name]['nickname'] = nickname
        users[name]['is_admin'] = is_admin
        users[name]['is_agent'] = is_agent
        users[name]['enabled'] = enabled
        if product is not None:
            users[name]['product'] = product
        users[name]['remark'] = remark
        save_users(users)
    return redirect(url_for('user_list'))


@app.route('/users/import', methods=['POST'])
@admin_required
def import_users():
    """
    批量导入用户（Excel）。
    用途：大批量账号导入。
    交互：导入成功自动写入台账。
    """
    file = request.files.get('file')
    price = float(request.form.get('price') or 0)
    product = request.form.get('product', '')
    if not file:
        return redirect(url_for('user_list'))
    filename = secure_filename(file.filename)
    if not filename:
        return redirect(url_for('user_list'))
    wb = load_workbook(file)
    ws = wb.active
    users = load_users()
    first = True
    count = 0
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        username = str(row[0]) if row and row[0] else None
        password = str(row[1]) if row and len(row) > 1 else None
        nickname = str(row[2]) if row and len(row) > 2 else ''
        is_admin = bool(row[3]) if row and len(row) > 3 else False
        if username and password:
            users[username] = {
                'user_id': generate_user_id(users),
                'password': password,
                'nickname': nickname,
                'is_admin': is_admin,
                'enabled': True,
                'source': 'import',
                'product': product,
                'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'last_login': None,
                'price': price,
                'ip_address': '',
                'location': ''
            }
            count += 1
    save_users(users)
    if count > 0 and price > 0:
        records = load_ledger()
        records.append({
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'admin': session.get('admin'),
            'role': 'admin',
            'product': product,
            'price': price,
            'count': count,
            'revenue': price * count
        })
        save_ledger(records)
    return redirect(url_for('user_list'))


@app.route('/users/export')
@admin_required
def export_users():
    """
    导出所有用户信息为Excel文件。
    用途：管理员备份、分析。
    """
    users = load_users()
    wb = Workbook()
    ws = wb.active
    ws.append([
        '用户编号', '用户名', '密码', '昵称', '是否管理员',
        '启用', '来源', '创建时间', '最后登录', '产品', 'IP地址', '位置'
    ])
    for name, info in users.items():
        ws.append([
            info.get('user_id', ''),
            name,
            info.get('password'),
            info.get('nickname'),
            info.get('is_admin'),
            info.get('enabled'),
            info.get('source'),
            info.get('created_at'),
            info.get('last_login'),
            info.get('product',''),
            info.get('ip_address', ''),
            info.get('location', '')
        ])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, download_name='users_export.xlsx', as_attachment=True)


@app.route('/users/template')
@admin_required
def download_template():
    """
    下载用户导入模板Excel文件。
    用途：批量导入格式参考。
    """
    wb = Workbook()
    ws = wb.active
    ws.append(['用户名', '密码', '昵称', '是否管理员'])
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, download_name='import_template.xlsx', as_attachment=True)


@app.route('/products')
@admin_required
def products():
    """
    产品管理页面。
    用途：显示所有产品信息。
    """
    products = load_products()
    return render_template('products.html', products=products)


@app.route('/products/add', methods=['POST'])
@admin_required
def add_product():
    """
    添加新产品。
    用途：表单提交产品信息，若设置为默认则取消其它默认。
    """
    name = request.form.get('name')
    version = request.form.get('version', '')
    ptype = request.form.get('ptype', '')
    price = float(request.form.get('price') or 0)
    default = bool(request.form.get('default'))
    if not name:
        return redirect(url_for('products'))
    products = load_products()
    if default:
        for p in products.values():
            p['default'] = False
    products[name] = {
        'name': name,
        'version': version,
        'type': ptype,
        'price': price,
        'default': default
    }
    save_products(products)
    return redirect(url_for('products'))


@app.route('/products/<path:name>/delete', methods=['POST'])
@admin_required
def delete_product(name):
    """
    删除指定产品。
    用途：产品维护。
    """
    products = load_products()
    if name in products:
        products.pop(name)
        save_products(products)
    return redirect(url_for('products'))


@app.route('/products/<path:name>/default', methods=['POST'])
@admin_required
def set_default_product(name):
    """
    设置指定产品为默认产品。
    用途：下单时默认选择。
    """
    products = load_products()
    if name in products:
        for p in products.values():
            p['default'] = False
        products[name]['default'] = True
        save_products(products)
    return redirect(url_for('products'))


@app.route('/users/bulk_create', methods=['POST'])
@admin_required
def bulk_create():
    """
    批量创建随机新用户。
    用途：管理员大批量生成账号。
    交互：会话保存本次批量信息，写入台账。
    """
    count = int(request.form.get('count', 0))
    price = float(request.form.get('price') or 0)
    product = request.form.get('product', '')
    users = load_users()
    new_accounts = []
    for _ in range(count):
        while True:
            uname = f"huiying{os.urandom(4).hex()}"[:12]
            if uname not in users:
                break
        pwd = os.urandom(4).hex()
        users[uname] = {
            'user_id': generate_user_id(users),
            'password': pwd,
            'nickname': '',
            'is_admin': False,
            'enabled': True,
            'source': 'batch',
            'product': product,
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'last_login': None,
            'price': price,
            'ip_address': '',
            'location': ''
        }
        new_accounts.append({'username': uname, 'password': pwd})
    save_users(users)
    session['bulk_accounts'] = new_accounts
    session['bulk_info'] = {
        'product': product,
        'price': price,
        'admin': session.get('admin'),
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
    if count > 0 and price > 0:
        records = load_ledger()
        records.append({
            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'admin': session.get('admin'),
            'product': product,
            'price': price,
            'count': count,
            'revenue': price * count
        })
        save_ledger(records)
    return redirect(url_for('bulk_manage'))


@app.route('/ledger')
@admin_required
def ledger_view():
    """
    台账页面（仅管理员）。
    用途：显示收入统计、筛选、导出。
    交互：仅统计role=admin的记录，避免重复计算。
    """
    records = load_ledger()
    product_filter = request.args.get('product', '')
    admin_filter = request.args.get('admin', '')
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    
    # 过滤记录
    # Only show admin role records to avoid counting agent sales
    filtered_records = [r for r in records if r.get('role') == 'admin']
    if product_filter:
        filtered_records = [r for r in filtered_records if r.get('product') == product_filter]
    if admin_filter:
        filtered_records = [r for r in filtered_records if r.get('admin') == admin_filter]
    if start:
        filtered_records = [r for r in filtered_records if r.get('time', '') >= start]
    if end:
        filtered_records = [r for r in filtered_records if r.get('time', '') <= end]
    
    # 计算统计数据
    from datetime import datetime
    today = datetime.now().strftime('%Y-%m-%d')
    this_month = datetime.now().strftime('%Y-%m')
    this_year = datetime.now().strftime('%Y')
    
    # 计算各时间段收入
    daily = sum(
        r.get('revenue', 0) for r in records
        if r.get('role') == 'admin' and r.get('time', '').startswith(today)
    )
    monthly = sum(
        r.get('revenue', 0) for r in records
        if r.get('role') == 'admin' and r.get('time', '').startswith(this_month)
    )
    yearly = sum(
        r.get('revenue', 0) for r in records
        if r.get('role') == 'admin' and r.get('time', '').startswith(this_year)
    )
    total = sum(r.get('revenue', 0) for r in records if r.get('role') == 'admin')
    
    products = load_products()
    return render_template(
        'ledger.html', records=filtered_records,
        product_filter=product_filter, admin_filter=admin_filter,
        start=start, end=end, products=products,
        daily=daily, monthly=monthly, yearly=yearly, total=total
    )


@app.route('/sales/users')
@agent_required
def agent_users():
    """
    代理名下用户管理页面。
    用途：代理自助筛选、分页、排序、管理账号。
    """
    users = load_users()
    current = session.get('agent')
    my_users = {k: v for k, v in users.items() if v.get('owner') == current}

    query = request.args.get('q', '')
    sale = request.args.get('sale', '')
    status = request.args.get('status', '')
    sort = request.args.get('sort', 'desc')
    start = request.args.get('start', '')
    end = request.args.get('end', '')
    page = int(request.args.get('page', 1))
    per_page = max(int(request.args.get('per_page', 20)), 1)

    if query:
        my_users = {k: v for k, v in my_users.items() if query.lower() in k.lower()}
    if status:
        flag = status == 'enabled'
        my_users = {k: v for k, v in my_users.items() if v.get('enabled', True) == flag}
    if sale:
        flag = sale == 'forsale'
        my_users = {k: v for k, v in my_users.items() if v.get('forsale', False) == flag}
    if start:
        my_users = {k: v for k, v in my_users.items() if v.get('created_at', '') >= start}
    if end:
        my_users = {k: v for k, v in my_users.items() if v.get('created_at', '') <= end}

    items = list(my_users.items())
    items.sort(key=lambda x: x[1].get('created_at', ''), reverse=(sort != 'asc'))
    total = len(items)
    page_items = items[(page - 1) * per_page: page * per_page]

    return render_template(
        'users.html',
        users=dict(page_items),
        total=total,
        page=page,
        per_page=per_page,
        query=query,
        source='',
        status=status,
        sale=sale,
        sort=sort,
        start=start,
        end=end,
        products=load_products()
    )


@app.route('/sales/ledger')
@agent_required
def agent_ledger():
    """
    代理销售台账页面。
    用途：仅显示当前代理的销售记录与统计。
    """
    product_filter = request.args.get('product', '')
    start = request.args.get('start', '')
    end = request.args.get('end', '')

    records = [
        r for r in load_ledger()
        if r.get('admin') == session.get('agent') and r.get('role') == 'agent'
    ]
    if product_filter:
        records = [r for r in records if r.get('product') == product_filter]
    if start:
        records = [r for r in records if r.get('time', '') >= start]
    if end:
        records = [r for r in records if r.get('time', '') <= end]

    from datetime import datetime
    today = datetime.now().strftime('%Y-%m-%d')
    this_month = datetime.now().strftime('%Y-%m')
    this_year = datetime.now().strftime('%Y')

    daily = sum(r.get('revenue', 0) for r in records if r.get('time', '').startswith(today))
    monthly = sum(r.get('revenue', 0) for r in records if r.get('time', '').startswith(this_month))
    yearly = sum(r.get('revenue', 0) for r in records if r.get('time', '').startswith(this_year))
    total = sum(r.get('revenue', 0) for r in records)

    products = load_products()
    return render_template(
        'ledger.html', records=records,
        product_filter=product_filter, admin_filter=session.get('agent'),
        start=start, end=end, products=products,
        daily=daily, monthly=monthly, yearly=yearly, total=total
    )


@app.route('/sales/apply', methods=['GET', 'POST'])
@agent_required
def apply_bulk():
    """
    代理批量申请账号页面。
    用途：提交申请，显示申请历史。
    交互：写入applications.json，session标记提交成功。
    """
    products = load_products()
    if request.method == 'POST':
        count = int(request.form.get('count', 0))
        price = float(request.form.get('price') or 0)
        product = request.form.get('product', '')
        apps = load_applications()
        apps.append({
            'id': os.urandom(6).hex(),
            'agent': session.get('agent'),
            'count': count,
            'price': price,
            'product': product,
            'status': 'pending',
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        })
        save_applications(apps)
        session['apply_success'] = True
        return redirect(url_for('apply_bulk'))
    success = session.pop('apply_success', False)
    my_apps = [a for a in load_applications() if a.get('agent') == session.get('agent')]
    return render_template(
        'bulk.html', accounts=None, products=products, info=None,
        page=1, per_page=20, total=0, success=success, apps=my_apps
    )


@app.route('/applications')
@admin_required
def applications_list():
    """
    管理员审批页面，显示所有代理批量申请。
    """
    apps = load_applications()
    return render_template('applications.html', apps=apps, products=load_products())


def _approve_application(app_record):
    """
    内部函数：审批通过代理批量申请，批量生成账号并写入台账。
    用途：供审批接口调用。
    """
    users = load_users()
    new_accounts = []
    for _ in range(app_record['count']):
        while True:
            uname = f"huiying{os.urandom(4).hex()}"[:12]
            if uname not in users:
                break
        pwd = os.urandom(4).hex()
        users[uname] = {
            'user_id': generate_user_id(users),
            'password': pwd,
            'nickname': '',
            'is_admin': False,
            'enabled': True,
            'source': 'agent',
            'product': app_record['product'],
            'created_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'last_login': None,
            'price': app_record['price'],
            'ip_address': '',
            'location': '',
            'owner': app_record['agent'],
            'forsale': True
        }
        new_accounts.append({'username': uname, 'password': pwd})
    save_users(users)
    records = load_ledger()
    records.append({
        'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'admin': session.get('admin'),
        'agent': app_record['agent'],
        'role': 'admin',
        'product': app_record['product'],
        'price': app_record['price'],
        'count': app_record['count'],
        'revenue': app_record['price'] * app_record['count']
    })
    save_ledger(records)
    app_record['status'] = 'approved'


@app.route('/applications/<app_id>/approve', methods=['POST'])
@admin_required
def approve_application(app_id):
    """
    审批通过指定代理批量申请。
    用途：管理员操作。
    """
    apps = load_applications()
    app_record = next((a for a in apps if a['id'] == app_id), None)
    if not app_record or app_record['status'] != 'pending':
        return redirect(url_for('applications_list'))
    _approve_application(app_record)
    save_applications(apps)
    return redirect(url_for('applications_list'))


@app.route('/applications/<app_id>/reject', methods=['POST'])
@admin_required
def reject_application(app_id):
    """
    拒绝指定代理批量申请。
    用途：管理员操作。
    """
    apps = load_applications()
    for a in apps:
        if a['id'] == app_id and a['status'] == 'pending':
            a['status'] = 'rejected'
            break
    save_applications(apps)
    return redirect(url_for('applications_list'))


@app.route('/applications/<app_id>/update', methods=['POST'])
@admin_required
def update_application(app_id):
    """
    更新指定申请的数量、单价、产品。
    用途：管理员审批前可修正申请内容。
    """
    apps = load_applications()
    app_record = next((a for a in apps if a['id'] == app_id), None)
    if not app_record or app_record.get('status') != 'pending':
        return redirect(url_for('applications_list'))
    app_record['count'] = int(request.form.get('count', app_record['count']))
    app_record['price'] = float(request.form.get('price', app_record['price']))
    app_record['product'] = request.form.get('product', app_record['product'])
    save_applications(apps)
    return redirect(url_for('applications_list'))


@app.route('/applications/batch', methods=['POST'])
@admin_required
def batch_applications():
    """
    批量审批或拒绝多个代理申请。
    用途：提升审批效率。
    """
    action = request.form.get('action')
    ids = request.form.getlist('ids')
    if action == 'approve':
        apps = load_applications()
        for app_record in apps:
            if app_record['id'] in ids and app_record['status'] == 'pending':
                _approve_application(app_record)
        save_applications(apps)
    elif action == 'reject':
        apps = load_applications()
        for a in apps:
            if a['id'] in ids and a['status'] == 'pending':
                a['status'] = 'rejected'
        save_applications(apps)
    return redirect(url_for('applications_list'))



if __name__ == '__main__':
    # 主程序入口，支持命令行指定端口
    parser = argparse.ArgumentParser()
    parser.add_argument('--port', type=int, default=5001, help='Port to run the server on')
    args = parser.parse_args()
    app.run(host='0.0.0.0', port=args.port, debug=True)

